"""
IAS 19 – Severance Pay Liability Valuation System
Valuation Date: 31 December 2024
Method: Projected Unit Credit (IAS 19)
"""

import streamlit as st
import pandas as pd
import numpy as np
from datetime import date
import io, os
import openpyxl
import plotly.graph_objects as go
import plotly.express as px

# ──────────────────────────────────────────────────────────────
# CONSTANTS
# ──────────────────────────────────────────────────────────────
VALUATION_DATE    = date(2024, 12, 31)
RETIREMENT_AGE    = {"M": 67, "F": 64}
DEFAULT_SALARY_GROWTH = 0.05

# ──────────────────────────────────────────────────────────────
# DATA LOADING
# ──────────────────────────────────────────────────────────────

def load_employees(path) -> pd.DataFrame:
    # Read with pandas — handles both xlsx and CSV transparently, and returns
    # all values as proper Python types (numbers stay numeric, dates are parsed).
    df_raw = pd.read_excel(path, sheet_name=0, header=None)

    def to_date(v):
        if v is None or (isinstance(v, float) and np.isnan(v)): return None
        try:
            return pd.to_datetime(v).date()
        except Exception:
            return None

    records = []
    for _, row in df_raw.iloc[2:].iterrows():   # rows 0=totals, 1=headers
        gender = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ""
        if gender not in ("M", "F"):
            continue

        try:
            salary = float(row.iloc[6]) if pd.notna(row.iloc[6]) else 0.0
            pct_raw = _to_float(row.iloc[8])
            if pct_raw is None or (isinstance(pct_raw, float) and np.isnan(pct_raw)):
                sec14_pct = 0.0
            elif pct_raw > 1.0:
                sec14_pct = pct_raw / 100.0   # stored as whole number (e.g. 72 or 100)
            else:
                sec14_pct = pct_raw           # stored as fraction (e.g. 0.72 or 1.0)
            fund = float(row.iloc[9]) if pd.notna(row.iloc[9]) else 0.0
        except Exception:
            salary, sec14_pct, fund = 0.0, 0.0, 0.0

        if salary <= 0:
            continue

        start_date  = to_date(row.iloc[5])
        sec14_raw   = row.iloc[7]
        sec14_start = to_date(sec14_raw) if pd.notna(sec14_raw) else start_date
        if sec14_start is None:
            sec14_start = start_date

        records.append(dict(
            first_name    = str(row.iloc[1]) if pd.notna(row.iloc[1]) else "",
            last_name     = str(row.iloc[2]) if pd.notna(row.iloc[2]) else "",
            gender        = gender,
            dob           = to_date(row.iloc[4]),
            start_date    = start_date,
            salary        = salary,
            sec14_pct     = sec14_pct,
            sec14_start   = sec14_start,
            existing_fund = fund,
            end_date      = to_date(row.iloc[11]),
            departure_type= str(row.iloc[14]) if pd.notna(row.iloc[14]) else "",
        ))
    return pd.DataFrame(records)


def _to_float(v):
    """Convert cell value (number or numeric string) to float, else None."""
    if isinstance(v, (int, float)): return float(v)
    try: return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError): return None

def load_assumptions(path):
    wb   = openpyxl.load_workbook(path, data_only=True)
    ws   = wb.worksheets[1]
    rows = list(ws.iter_rows(values_only=True))

    def to_rate(val):
        """Normalise a rate that may be stored as fraction (0.05) or percent (5)."""
        if val is None: return None
        return val / 100.0 if val > 1.0 else val

    curve, turns, sg = {}, {}, DEFAULT_SALARY_GROWTH
    for row in rows[3:]:
        if not row: continue
        yr  = _to_float(row[1])
        rt  = to_rate(_to_float(row[2]))
        if yr is not None and rt is not None:
            curve[int(yr)] = rt
        ag  = row[5]
        fr  = to_rate(_to_float(row[6]))
        rr  = to_rate(_to_float(row[7]))
        if isinstance(ag, str) and "-" in ag and fr is not None:
            turns[ag] = {"fire": fr, "resign": rr or 0.0}
        sg_val = to_rate(_to_float(row[10]))
        if sg_val is not None and 0 < sg_val < 1:
            sg = sg_val
    return curve, turns, sg


def load_mortality(path) -> dict:
    mort = {"M": {}, "F": {}}
    base = os.path.dirname(os.path.abspath(path))

    # Try two separate CSV files first (male / female)
    csv_m = os.path.join(base, "mortality_male.csv")
    csv_f = os.path.join(base, "mortality_female.csv")
    if os.path.exists(csv_m) and os.path.exists(csv_f):
        for gender, csv_path in [("M", csv_m), ("F", csv_f)]:
            df = pd.read_csv(csv_path, header=None)
            for _, row in df.iterrows():
                age = _to_float(row.iloc[1])
                q   = _to_float(row.iloc[5]) if len(row) > 5 else None
                if age is not None and q is not None:
                    mort[gender][int(age)] = q
        return mort

    # Fall back to a two-sheet xlsx
    wb = openpyxl.load_workbook(path, data_only=True)
    if len(wb.worksheets) < 2:
        raise ValueError(
            "mortality_table.xlsx must contain two sheets (males first, females second), "
            "or place mortality_male.csv / mortality_female.csv alongside the workbook.")
    for gender, ws in zip(["M", "F"], wb.worksheets[:2]):
        for row in ws.iter_rows(values_only=True):
            age = _to_float(row[1])
            q   = _to_float(row[5]) if len(row) > 5 else None
            if age is not None and q is not None:
                mort[gender][int(age)] = q
    return mort


# ──────────────────────────────────────────────────────────────
# ACTUARIAL HELPERS
# ──────────────────────────────────────────────────────────────

def years_between(d1, d2) -> float:
    return (d2 - d1).days / 365.25

def turnover_for_age(age, rates):
    if   age <= 29: key = "18-29"
    elif age <= 39: key = "30-39"
    elif age <= 49: key = "40-49"
    elif age <= 59: key = "50-59"
    else:           key = "60-67"
    r = rates.get(key, {"fire": 0.02, "resign": 0.03})
    return r["fire"], r["resign"]

def mortality_for_age(age, gender, mort):
    return mort[gender].get(max(18, min(int(age), 110)), 0.0)

def discount_factor(t, curve):
    if t <= 0: return 1.0
    max_y = max(curve)
    t_eff = min(t, max_y)
    if t_eff <= 1:
        r = curve.get(1, curve[min(curve.keys())])
    else:
        t_lo = int(t_eff)
        t_hi = min(t_lo + 1, max_y)
        r_lo = curve.get(t_lo, curve[max_y])
        r_hi = curve.get(t_hi, curve[max_y])
        r    = r_lo + (t_eff - t_lo) * (r_hi - r_lo)
    if t > max_y:
        r = curve[max_y]
    return (1 + r) ** (-t)


# ──────────────────────────────────────────────────────────────
# DBO CALCULATION  (Projected Unit Credit)
# ──────────────────────────────────────────────────────────────

def calculate_dbo(emp: dict, curve, turns, sg, mort):
    """
    Returns (gross_DBO, net_liability, detail_rows).
    IAS 19 Projected Unit Credit — prorated benefit formula.
    """
    gender = emp["gender"]; dob = emp["dob"]; start = emp["start_date"]
    salary = emp["salary"]; sec14 = emp["sec14_pct"]; fund = emp["existing_fund"]

    W = RETIREMENT_AGE.get(gender, 67)
    x = years_between(dob, VALUATION_DATE)
    v = years_between(start, VALUATION_DATE)  # seniority at valuation date
    n = int(W - x)

    if n <= 0:
        b = salary * max(v, 0) * (1 - sec14)
        row = dict(year=0, age=round(x,1), seniority=round(v,2),
                   salary_proj=round(salary,2), lp=1.0,
                   q_dismissal=0.0, q_resign=0.0, q_death=0.0,
                   benefit=round(b,2), disc_factor=1.0, prorate=1.0,
                   dbo_dismissal=0.0, dbo_resign=0.0,
                   dbo_death=0.0, dbo_retirement=round(b,2),
                   year_dbo=round(b,2), cum_dbo=round(b,2))
        return round(b, 2), round(max(0.0, b - fund), 2), [row]

    lp, dbo, detail, cum = 1.0, 0.0, [], 0.0

    for t in range(1, n + 1):
        age_t    = int(x) + t - 1
        sen_exit = v + t   # total seniority at projected exit
        sal_t    = salary * (1 + sg) ** t

        # א. Projected benefit — current seniority v (prorating embedded in formula)
        benefit = sal_t * v * (1 - sec14)

        # ב. Exit probabilities
        q_f, q_r = turnover_for_age(age_t, turns)
        q_d      = mortality_for_age(age_t, gender, mort)

        # ג. Employer payment probability
        q_pay = q_f + q_d                          # dismissal + death — always liable
        if sec14 == 0 and (v + t) >= 5:
            q_pay += q_r                           # resignation — only if no Sec14 & sen ≥ 5
        q_pay_r = q_r if (sec14 == 0 and (v + t) >= 5) else 0.0

        # ד. Present value
        disc_mid = discount_factor(t - 0.5, curve)
        year_dbo = lp * q_pay * benefit * disc_mid

        # ה. Retirement (final year)
        if t == n:
            lp_ret    = lp * (1 - (q_f + q_r + q_d))
            year_dbo += lp_ret * benefit * discount_factor(n, curve)

        dbo += year_dbo; cum += year_dbo

        detail.append(dict(
            year=t, age=age_t, seniority=round(sen_exit, 2),
            salary_proj=round(sal_t, 2), lp=round(lp, 5),
            q_dismissal=round(q_f, 4), q_resign=round(q_r, 4), q_death=round(q_d, 6),
            benefit=round(benefit, 2), disc_factor=round(disc_mid, 5),
            prorate=round(v / sen_exit, 4),
            dbo_dismissal=round(lp * q_f  * benefit * disc_mid, 2),
            dbo_resign   =round(lp * q_pay_r * benefit * disc_mid, 2),
            dbo_death    =round(lp * q_d  * benefit * disc_mid, 2),
            dbo_retirement=round(year_dbo - lp * q_pay * benefit * disc_mid, 2),
            year_dbo=round(year_dbo, 2), cum_dbo=round(cum, 2),
        ))
        lp *= (1 - (q_f + q_r + q_d))

    return round(dbo, 2), round(max(0.0, dbo - fund), 2), detail


# ──────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ──────────────────────────────────────────────────────────────

def export_to_excel(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "IAS19_Liability"
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append([v if not isinstance(v, bool) else str(v) for v in row])
    # Basic number formatting
    from openpyxl.styles import Font, Alignment, PatternFill
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 28)
    wb.save(buf)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────
# UI HELPERS
# ──────────────────────────────────────────────────────────────

CSS = """
<style>
  [data-testid="stAppViewContainer"] { background: #f1f5f9; }
  [data-testid="stSidebar"]          { background: #1e3a5f; }
  [data-testid="stSidebar"] *        { color: #e2eaf5 !important; }

  .kpi-row { display:flex; gap:1rem; margin:0.8rem 0 1.2rem; }
  .kpi {
    flex:1; background:white; border-radius:10px;
    padding:1rem 1.2rem; box-shadow:0 1px 6px rgba(0,0,0,0.08);
    border-left:4px solid #2563eb;
  }
  .kpi.red  { border-left-color:#dc2626; }
  .kpi.teal { border-left-color:#0891b2; }
  .kpi.violet { border-left-color:#7c3aed; }
  .kpi-label { font-size:.72rem; font-weight:700; color:#64748b;
               text-transform:uppercase; letter-spacing:.06em; }
  .kpi-value { font-size:1.65rem; font-weight:800; color:#1e293b; margin:.2rem 0; }
  .kpi-sub   { font-size:.72rem; color:#94a3b8; }

  .section-title {
    font-size:1rem; font-weight:700; color:#1e3a5f;
    border-bottom:2px solid #dbeafe; padding-bottom:.4rem;
    margin:1.2rem 0 .7rem;
  }
  .info-grid {
    display:grid; grid-template-columns:1fr 1fr; gap:.5rem 2rem;
    background:white; border-radius:10px; padding:1.1rem 1.4rem;
    box-shadow:0 1px 6px rgba(0,0,0,0.07); margin-bottom:1rem;
  }
  .info-row { display:flex; justify-content:space-between;
              padding:.25rem 0; border-bottom:1px solid #f1f5f9; }
  .info-label { font-size:.8rem; color:#64748b; font-weight:600; }
  .info-value { font-size:.85rem; color:#1e293b; font-weight:700; text-align:right; }

  .crit-table { width:100%; border-collapse:collapse; }
  .crit-table td { padding:.35rem .5rem; font-size:.82rem;
                   border-bottom:1px solid #f1f5f9; }
  .pass  { color:#16a34a; font-weight:700; }
  .fail  { color:#dc2626; font-weight:700; }
  .info  { color:#2563eb; font-weight:700; }
  .warn  { color:#d97706; font-weight:700; }

  .result-banner {
    background: linear-gradient(135deg,#1e3a5f,#1d4ed8);
    border-radius:10px; padding:1rem 1.5rem; color:white;
    display:flex; justify-content:space-around; margin:1rem 0;
    box-shadow:0 2px 8px rgba(30,58,95,.25);
  }
  .rb-item { text-align:center; }
  .rb-label { font-size:.72rem; opacity:.75; font-weight:600;
              text-transform:uppercase; letter-spacing:.06em; }
  .rb-value { font-size:1.5rem; font-weight:800; }
</style>
"""


def kpi_html(label, value, sub="", cls=""):
    return (f'<div class="kpi {cls}">'
            f'<div class="kpi-label">{label}</div>'
            f'<div class="kpi-value">{value}</div>'
            f'<div class="kpi-sub">{sub}</div>'
            f'</div>')


def render_criteria(emp, x, v, n, gross, net):
    W     = RETIREMENT_AGE.get(emp["gender"], 67)
    active = emp["end_date"] is None or emp["end_date"] > VALUATION_DATE
    left24 = emp["end_date"] is not None and emp["end_date"].year == 2024
    sec14_lbl = (f"{emp['sec14_pct']*100:.0f}%"
                 + (" (full coverage — no liability)" if emp["sec14_pct"] == 1.0
                    else f" (employer bears {(1-emp['sec14_pct'])*100:.0f}%)"))
    retire_yr = VALUATION_DATE.year + max(n, 0)

    rows = [
        ("✓", "pass",  "Active on 31 Dec 2024",
         "Active" if active else "Departed"),
        ("✓" if not left24 else "✗", "pass" if not left24 else "fail",
         "Did not leave during 2024",
         "Qualifies" if not left24 else "Left in 2024 → Liability = 0"),
        ("✓" if v >= 1 else "⚠", "pass" if v >= 1 else "warn",
         "Minimum seniority for severance (≥ 1 yr)",
         f"Seniority: {v:.1f} yrs"),
        ("✓" if n > 0 else "⚠", "pass" if n > 0 else "warn",
         "Has not yet reached retirement age",
         f"Retirement age: {W}  |  Years remaining: {max(n,0)}"),
        ("i", "info",  "Section 14 coverage",  sec14_lbl),
        ("i", "info",  "Expected retirement year", str(retire_yr)),
        ("i", "info",  "Gross DBO (before fund deduction)", f"₪ {gross:,.0f}"),
        ("i", "info",  "Net Liability (IAS 19)", f"₪ {net:,.0f}"),
    ]
    html = ('<div style="background:white;border-radius:10px;padding:.8rem 1.1rem;'
            'box-shadow:0 1px 6px rgba(0,0,0,.07)">'
            '<table class="crit-table">')
    for icon, cls, label, note in rows:
        html += (f"<tr>"
                 f"<td style='width:1.5rem'><span class='{cls}'>{icon}</span></td>"
                 f"<td style='color:#374151;font-weight:600'>{label}</td>"
                 f"<td style='color:#6b7280'>{note}</td>"
                 f"</tr>")
    html += "</table></div>"
    return html


def render_result_banner(gross, fund, net):
    return (f'<div class="result-banner">'
            f'<div class="rb-item"><div class="rb-label">Gross DBO</div>'
            f'<div class="rb-value">₪ {gross:,.0f}</div></div>'
            f'<div class="rb-item" style="opacity:.4">→</div>'
            f'<div class="rb-item"><div class="rb-label">Existing Fund</div>'
            f'<div class="rb-value">₪ {fund:,.0f}</div></div>'
            f'<div class="rb-item" style="opacity:.4">=</div>'
            f'<div class="rb-item"><div class="rb-label">Net Liability</div>'
            f'<div class="rb-value">₪ {net:,.0f}</div></div>'
            f'</div>')


# ──────────────────────────────────────────────────────────────
# MAIN APP
# ──────────────────────────────────────────────────────────────

def main():
    st.set_page_config(
        page_title="IAS 19 – Severance Liability",
        page_icon="📊",
        layout="wide",
        initial_sidebar_state="collapsed",
    )
    st.markdown(CSS, unsafe_allow_html=True)

    # ── Header ───────────────────────────────────────────────
    st.markdown("""
    <div style="background:linear-gradient(135deg,#1e3a5f 0%,#1d4ed8 100%);
         border-radius:12px;padding:1.4rem 2rem;margin-bottom:1.2rem">
      <h2 style="color:white;margin:0;font-size:1.5rem">
        IAS 19 — Severance Pay Liability Valuation
      </h2>
      <p style="color:#93c5fd;margin:.3rem 0 0;font-size:.88rem">
        Projected Unit Credit Method &nbsp;|&nbsp;
        Valuation Date: <strong style="color:white">31 December 2024</strong>
      </p>
    </div>
    """, unsafe_allow_html=True)

    # ── Load files ────────────────────────────────────────────
    base      = os.path.dirname(os.path.abspath(__file__))
    data_path = os.path.join(base, "data10.xlsx")
    mort_path = os.path.join(base, "mortality_table.xlsx")

    for p, name in [(data_path, "data10.xlsx"), (mort_path, "mortality_table.xlsx")]:
        if not os.path.exists(p):
            st.error(f"File not found: **{name}**  (expected at `{p}`)")
            st.stop()

    @st.cache_data
    def load_all():
        emps             = load_employees(data_path)
        curve, turns, sg = load_assumptions(data_path)
        mort             = load_mortality(mort_path)
        return emps, curve, turns, sg, mort

    with st.spinner("Loading files and running actuarial calculations…"):
        df, curve, turns, sg, mort = load_all()

    # ── Derive status fields ──────────────────────────────────
    df["age"]         = df["dob"].apply(lambda d: round(years_between(d, VALUATION_DATE), 1))
    df["seniority"]   = df["start_date"].apply(lambda d: round(years_between(d, VALUATION_DATE), 1))
    df["active"]      = df["end_date"].apply(lambda d: d is None or d > VALUATION_DATE)
    df["left_2024"]   = df["end_date"].apply(lambda d: d is not None and d.year == 2024)

    # ── Run DBO for every employee ─────────────────────────────
    gross_list, net_list = [], []
    for _, emp in df.iterrows():
        if not emp["active"] or emp["left_2024"]:
            gross_list.append(0.0); net_list.append(0.0)
        else:
            g, n, _ = calculate_dbo(emp.to_dict(), curve, turns, sg, mort)
            gross_list.append(g); net_list.append(n)
    df["gross_DBO"]     = gross_list
    df["net_liability"] = net_list

    # ── KPI cards ─────────────────────────────────────────────
    c1, c2, c3, c4 = st.columns(4)
    c1.markdown(
        '<div class="kpi-row">' +
        kpi_html("Total Employees", f"{len(df)}",
                 f"Active: {int(df['active'].sum())}") +
        '</div>', unsafe_allow_html=True)
    c2.markdown(
        '<div class="kpi-row">' +
        kpi_html("Gross DBO", f"₪ {df['gross_DBO'].sum():,.0f}",
                 "Before fund deduction", "violet") +
        '</div>', unsafe_allow_html=True)
    c3.markdown(
        '<div class="kpi-row">' +
        kpi_html("Plan Assets (Fund)", f"₪ {df['existing_fund'].sum():,.0f}",
                 "Existing severance fund", "teal") +
        '</div>', unsafe_allow_html=True)
    c4.markdown(
        '<div class="kpi-row">' +
        kpi_html("Net Liability", f"₪ {df['net_liability'].sum():,.0f}",
                 "IAS 19 Balance Sheet Amount", "red") +
        '</div>', unsafe_allow_html=True)

    # ── Tabs ──────────────────────────────────────────────────
    tab_list, tab_detail, tab_charts = st.tabs([
        "👥  Employee List",
        "🔍  Calculation Detail",
        "📊  Charts",
    ])

    # ══════════════════════════════════════════════════════════
    # TAB 1 – Employee list  (clickable rows)
    # ══════════════════════════════════════════════════════════
    with tab_list:
        st.markdown(
            '<div class="section-title">Employee Summary — click any row to view full calculation</div>',
            unsafe_allow_html=True)

        def status_label(row):
            if row["left_2024"]:  return "Left 2024"
            if not row["active"]: return "Departed"
            return "Active"

        def sec14_label(p):
            if p == 1.0:       return "100% – Full"
            if round(p,2) == 0.72: return "72% – Partial"
            if p > 0:          return f"{p*100:.0f}% – Partial"
            return "None"

        table_df = pd.DataFrame({
            "#":                  range(1, len(df)+1),
            "First Name":         df["first_name"],
            "Last Name":          df["last_name"],
            "Gender":             df["gender"].map({"M": "Male", "F": "Female"}),
            "Age":                df["age"],
            "Seniority (yrs)":    df["seniority"],
            "Monthly Salary":     df["salary"].round(0).astype(int),
            "Section 14":         df["sec14_pct"].apply(sec14_label),
            "Existing Fund":      df["existing_fund"].round(0).astype(int),
            "Status":             df.apply(status_label, axis=1),
            "Gross DBO":          df["gross_DBO"].round(0).astype(int),
            "Net Liability":      df["net_liability"].round(0).astype(int),
        })

        selection = st.dataframe(
            table_df,
            on_select="rerun",
            selection_mode="single-row",
            use_container_width=True,
            height=500,
            column_config={
                "#":               st.column_config.NumberColumn(width="small"),
                "Age":             st.column_config.NumberColumn(format="%.1f", width="small"),
                "Seniority (yrs)": st.column_config.NumberColumn(format="%.1f"),
                "Monthly Salary":  st.column_config.NumberColumn(format="₪%d"),
                "Existing Fund":   st.column_config.NumberColumn(format="₪%d"),
                "Gross DBO":       st.column_config.NumberColumn(format="₪%d"),
                "Net Liability":   st.column_config.NumberColumn(format="₪%d"),
            },
            hide_index=True,
        )

        if selection.selection.rows:
            selected_idx = selection.selection.rows[0]
            st.session_state["selected_emp"] = selected_idx
            st.info(
                f"✔ Row selected: **{table_df.iloc[selected_idx]['First Name']} "
                f"{table_df.iloc[selected_idx]['Last Name']}** — "
                f"switch to the **Calculation Detail** tab to view full breakdown.")

        # Download
        st.markdown("<br>", unsafe_allow_html=True)
        st.download_button(
            "⬇  Download Full Results (Excel)",
            data=export_to_excel(table_df),
            file_name="IAS19_Severance_31122024.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    # ══════════════════════════════════════════════════════════
    # TAB 2 – Calculation detail
    # ══════════════════════════════════════════════════════════
    with tab_detail:

        # ── Employee selector ────────────────────────────────
        default_idx = st.session_state.get("selected_emp", 0)
        emp_labels  = [
            f"{i+1}.  {r['first_name']} {r['last_name']}  "
            f"({'Male' if r['gender']=='M' else 'Female'}"
            f" | Age {df.loc[i,'age']} | Seniority {df.loc[i,'seniority']} yrs)"
            for i, r in df.iterrows()
        ]
        chosen = st.selectbox("Select employee:", options=emp_labels, index=default_idx)
        idx    = emp_labels.index(chosen)
        emp    = df.iloc[idx].to_dict()

        x = years_between(emp["dob"],        VALUATION_DATE)
        v = years_between(emp["start_date"], VALUATION_DATE)
        W = RETIREMENT_AGE.get(emp["gender"], 67)
        n = int(W - x)
        retire_yr = VALUATION_DATE.year + max(n, 0)
        gross, net, detail_rows = calculate_dbo(emp, curve, turns, sg, mort)

        st.markdown("---")

        # ── SECTION 1: Profile + Eligibility ────────────────
        col_prof, col_elig = st.columns(2, gap="large")

        with col_prof:
            st.markdown('<div class="section-title">Employee Profile</div>',
                        unsafe_allow_html=True)
            st.markdown(f"""
            <div style="background:white;border-radius:10px;padding:1rem 1.3rem;
                        box-shadow:0 1px 6px rgba(0,0,0,.07)">
              <table style="width:100%;border-collapse:collapse">
                {"".join(
                    f"<tr style='border-bottom:1px solid #f1f5f9'>"
                    f"<td style='padding:.35rem .5rem;color:#64748b;font-size:.82rem;font-weight:600;width:45%'>{lbl}</td>"
                    f"<td style='padding:.35rem .5rem;color:#1e293b;font-size:.85rem;font-weight:700'>{val}</td>"
                    f"</tr>"
                    for lbl, val in [
                        ("Full Name",            f"{emp['first_name']} {emp['last_name']}"),
                        ("Gender",               "Male" if emp["gender"]=="M" else "Female"),
                        ("Date of Birth",        emp["dob"].strftime("%d %b %Y")),
                        ("Age at 31 Dec 2024",   f"{x:.2f} years"),
                        ("Employment Start Date",emp["start_date"].strftime("%d %b %Y")),
                        ("Seniority at 31 Dec 2024", f"{v:.2f} years"),
                        ("Monthly Salary",       f"₪ {emp['salary']:,.0f}"),
                        ("Section 14 Coverage",  f"{emp['sec14_pct']*100:.0f}%"),
                        ("Existing Fund Balance",f"₪ {emp['existing_fund']:,.0f}"),
                        ("Retirement Age",       f"{W}  →  expected year {retire_yr}"),
                    ]
                )}
              </table>
            </div>
            """, unsafe_allow_html=True)

        with col_elig:
            st.markdown('<div class="section-title">Eligibility & Criteria Check</div>',
                        unsafe_allow_html=True)
            st.markdown(render_criteria(emp, x, v, n, gross, net), unsafe_allow_html=True)

        # ── SECTION 2: How the DBO is calculated ────────────
        st.markdown("---")
        st.markdown('<div class="section-title">How the DBO is Calculated — Step by Step</div>',
                    unsafe_allow_html=True)

        # Pick year 1 as the worked example
        ex_yr   = VALUATION_DATE.year + 1
        ex_sal  = emp["salary"] * (1 + sg) ** 1
        ex_sen  = v + 1
        ex_sec14_start = emp.get("sec14_start") or emp["start_date"]
        ex_vpre = max(0.0, years_between(emp["start_date"], ex_sec14_start))
        ex_spre = min(ex_vpre, ex_sen)
        ex_spost = max(0.0, ex_sen - ex_spre)
        ex_ben  = ex_sal * (ex_spre + ex_spost * (1.0 - emp["sec14_pct"]))
        ex_prorate = min(1.0, v / ex_sen) if ex_sen > 0 else 1.0
        q_f1, q_r1 = turnover_for_age(int(x), turns)
        q_d1       = mortality_for_age(int(x), emp["gender"], mort)
        q_tot1     = q_f1 + q_r1 + q_d1
        disc1      = discount_factor(0.5, curve)
        ex_fund1 = emp["existing_fund"] + (emp["sec14_pct"] * emp["salary"] * (1 + sg) * ((1 + sg) - 1) / sg if sg != 0 else emp["sec14_pct"] * emp["salary"] * 1)
        ex_resign1 = max(ex_ben, ex_fund1) if ex_sen >= 5 else 0.0
        ex_dbo     = 1.0 * (q_f1 * ex_ben + q_r1 * ex_resign1 + q_d1 * ex_ben) * disc1 * ex_prorate

        if emp["sec14_pct"] > 0 and ex_vpre > 0:
            sec14_note = (f"pre-Sec14 {ex_spre:.2f} yrs × 100% + "
                          f"post-Sec14 {ex_spost:.2f} yrs × {(1-emp['sec14_pct'])*100:.0f}%")
        elif emp["sec14_pct"] > 0:
            sec14_note = f"all service under Sec14 → {(1-emp['sec14_pct'])*100:.0f}% employer share"
        else:
            sec14_note = "no Section 14, employer bears 100%"

        st.markdown(f"""
        <div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:10px;
                    padding:1.2rem 1.5rem;font-size:.85rem;line-height:1.7">

          <p style="margin:0 0 .8rem;font-weight:700;color:#1e3a5f;font-size:.95rem">
            The IAS 19 Projected Unit Credit (PUC) method asks:<br>
            <em style="color:#2563eb">"For each future year, what is the expected present value
            of the severance benefit the employer may have to pay?"</em>
          </p>

          <p style="margin:0 0 .5rem"><strong>The formula applied to every projection year t:</strong></p>
          <div style="background:white;border-radius:8px;padding:.7rem 1rem;
                      font-family:monospace;color:#1e293b;margin-bottom:.8rem">
            DBO(t) = P(employed at t) &times; Exit Rate(t) &times; Benefit(t) &times; Discount Factor(t) &times; Prorate(t)
          </div>

          <p style="margin:0 0 .4rem;font-weight:700;color:#1e3a5f">
            Where for each year:
          </p>
          <ul style="margin:.2rem 0 .8rem 1.2rem;padding:0">
            <li><strong>Benefit(t)</strong> = Projected Salary &times;
                (pre-Sec14 years &times; 100% + post-Sec14 years &times; (1 &minus; Sec14%))<br>
                <span style="color:#64748b">Salary grows at {sg*100:.0f}%/yr; pre/post split determined by Sec14 start date</span></li>
            <li><strong>Prorate(t)</strong> = seniority at 31.12.24 &divide; seniority at exit in year t
                &nbsp;<em>(IAS 19 §67 — only past-service obligation is recognised)</em></li>
            <li><strong>Exit Rate(t)</strong> = dismissal + resignation* + mortality<br>
                <span style="color:#64748b">*resignation creates employer liability of max(legal benefit, fund value) for seniority &ge; 5 yrs (company policy)</span></li>
            <li><strong>Discount Factor(t)</strong> = present value of ₪1 due in t years
                (from the IAS 19 bond yield curve)</li>
            <li><strong>P(employed at t)</strong> = probability the employee has not left before year t</li>
          </ul>

          <p style="margin:0 0 .5rem;font-weight:700;color:#1e3a5f">
            Worked example — Year {ex_yr} (first projection year):
          </p>
          <table style="width:100%;border-collapse:collapse;background:white;
                        border-radius:8px;overflow:hidden">
            {"".join(
                f"<tr style='border-bottom:1px solid #f1f5f9'>"
                f"<td style='padding:.3rem .8rem;color:#64748b;font-size:.82rem;width:40%'>{lbl}</td>"
                f"<td style='padding:.3rem .8rem;color:#1e293b;font-weight:600;font-size:.82rem'>{val}</td>"
                f"</tr>"
                for lbl, val in [
                    ("Projected salary in 2025",
                     f"₪{emp['salary']:,.0f} × (1 + {sg*100:.0f}%)¹ = ₪{ex_sal:,.0f}"),
                    ("Seniority at exit (end of 2025)",
                     f"{v:.2f} + 1 = {ex_sen:.2f} years"),
                    ("Benefit split (Sec14 basis)",
                     sec14_note + f" → ₪{ex_ben:,.0f}"),
                    ("Prorate — IAS 19 §67",
                     f"v / t = {v:.2f} / {ex_sen:.2f} = {ex_prorate:.4f}"),
                    ("P(still employed at start of 2025)", "1.0000  (100% — no exits yet)"),
                    ("Exit rate in 2025",
                     f"Dismissal {q_f1:.1%} + Resignation {q_r1:.1%} + Death {q_d1:.4%} = {q_tot1:.4%}"),
                    ("Discount factor (mid-year, t = 0.5)",
                     f"{disc1:.5f}"),
                    ("DBO contribution from 2025",
                     f"1.0000 × (q_f×benefit + q_r×max(benefit,fund) + q_d×benefit) × {disc1:.5f} × {ex_prorate:.4f} ≈ ₪{ex_dbo:,.0f}"),
                ]
            )}
          </table>
          <p style="margin:.8rem 0 0;color:#64748b;font-size:.8rem">
            This is repeated for every year from {VALUATION_DATE.year+1} to {retire_yr}.
            The retirement year adds the benefit for the remaining probability of
            reaching age {W} with certainty.
            All contributions are summed → <strong>Gross DBO</strong>.
          </p>
        </div>
        """, unsafe_allow_html=True)

        # ── SECTION 3: Projection table ──────────────────────
        st.markdown("---")
        st.markdown('<div class="section-title">Full Year-by-Year Projection Table</div>',
                    unsafe_allow_html=True)

        if not detail_rows:
            st.info("No projection years — employee is at or past retirement age.")
        else:
            cd = pd.DataFrame(detail_rows)

            # Build simplified display table
            proj_df = pd.DataFrame({
                "Calendar Year":         VALUATION_DATE.year + cd["year"],
                "Age":                   cd["age"],
                "Seniority at Exit":     cd["seniority"],
                "Projected Salary (₪)":  cd["salary_proj"].map("{:,.0f}".format),
                "Benefit if Exit (₪)":   cd["benefit"].map("{:,.0f}".format),
                "Prorate (v/t)":         cd["prorate"].map("{:.4f}".format),
                "P(Employed)":           cd["lp"].map("{:.4f}".format),
                "Exit Rate":             (cd["q_dismissal"]+cd["q_resign"]+cd["q_death"]).map("{:.4f}".format),
                "Discount Factor":       cd["disc_factor"].map("{:.4f}".format),
                "DBO This Year (₪)":     cd["year_dbo"].map("{:,.0f}".format),
                "Cumulative DBO (₪)":    cd["cum_dbo"].map("{:,.0f}".format),
                "Event":                 ["🎓 RETIREMENT" if i == len(cd)-1 else ""
                                          for i in range(len(cd))],
            })

            st.dataframe(proj_df, use_container_width=True,
                         height=min(60 + 37*len(proj_df), 500),
                         hide_index=True,
                         column_config={
                             "Calendar Year":       st.column_config.NumberColumn(width="small", format="%d"),
                             "Age":                 st.column_config.NumberColumn(width="small"),
                             "Seniority at Exit":   st.column_config.NumberColumn(format="%.2f"),
                         })

            # Column legend
            st.markdown("""
            <div style="background:#f8fafc;border-radius:8px;padding:.7rem 1rem;
                        font-size:.78rem;color:#64748b;margin-top:.5rem">
              <strong>Column guide:</strong> &nbsp;
              <strong>Seniority at Exit</strong> — years of service if employee leaves that year &nbsp;|&nbsp;
              <strong>Benefit if Exit</strong> — employer's severance obligation:
                pre-Sec14 service at 100% + post-Sec14 service at (1−Sec14%) &nbsp;|&nbsp;
              <strong>Prorate (v/t)</strong> — IAS 19 §67 past-service ratio: seniority at 31.12.24 ÷ seniority at exit &nbsp;|&nbsp;
              <strong>P(Employed)</strong> — probability employee is still working at the start of that year &nbsp;|&nbsp;
              <strong>Exit Rate</strong> — combined probability of leaving (dismissal + resignation + death) &nbsp;|&nbsp;
              <strong>Discount Factor</strong> — present-value weight for that year &nbsp;|&nbsp;
              <strong>DBO This Year</strong> — P(Employed) × Exit Rate × Benefit × Discount Factor × Prorate
            </div>
            """, unsafe_allow_html=True)

            # Expandable probability detail
            with st.expander("Show full probability breakdown (dismissal / resignation / death)"):
                prob_df = pd.DataFrame({
                    "Calendar Year":    VALUATION_DATE.year + cd["year"],
                    "Age":              cd["age"],
                    "P(Employed) lp":   cd["lp"].map("{:.5f}".format),
                    "q Dismissal":      cd["q_dismissal"].map("{:.4f}".format),
                    "q Resignation":    cd["q_resign"].map("{:.4f}".format),
                    "q Death":          cd["q_death"].map("{:.6f}".format),
                    "Total Exit Rate":  (cd["q_dismissal"]+cd["q_resign"]+cd["q_death"]).map("{:.5f}".format),
                    "DBO — Dismissal (₪)":  cd["dbo_dismissal"].map("{:,.0f}".format),
                    "DBO — Resignation (₪)":cd["dbo_resign"].map("{:,.0f}".format),
                    "DBO — Death (₪)":      cd["dbo_death"].map("{:,.0f}".format),
                    "DBO — Retirement (₪)": cd["dbo_retirement"].map("{:,.0f}".format),
                })
                st.dataframe(prob_df, use_container_width=True,
                             height=min(60 + 37*len(prob_df), 450),
                             hide_index=True,
                             column_config={
                                 "Calendar Year": st.column_config.NumberColumn(format="%d"),
                             })

            # ── DBO build-up chart ───────────────────────────
            st.markdown('<div class="section-title">DBO Contribution by Year</div>',
                        unsafe_allow_html=True)
            cal_years = VALUATION_DATE.year + cd["year"]
            fig = go.Figure()
            fig.add_bar(x=cal_years, y=cd["dbo_dismissal"],
                        name="Dismissal",  marker_color="#1d4ed8", opacity=0.9)
            fig.add_bar(x=cal_years, y=cd["dbo_resign"],
                        name="Resignation",marker_color="#0891b2", opacity=0.9)
            fig.add_bar(x=cal_years, y=cd["dbo_death"],
                        name="Death",      marker_color="#6d28d9", opacity=0.9)
            fig.add_bar(x=cal_years, y=cd["dbo_retirement"],
                        name="Retirement", marker_color="#059669", opacity=0.9)
            fig.add_scatter(x=cal_years, y=cd["cum_dbo"],
                            name="Cumulative DBO",
                            mode="lines+markers",
                            line=dict(color="#1e293b", width=2, dash="dot"),
                            marker=dict(size=4, color="#1e293b"))
            fig.update_layout(
                barmode="stack", height=330,
                xaxis_title="Calendar Year",
                yaxis_title="₪",
                legend=dict(orientation="h", y=1.13, x=0, font=dict(size=11)),
                plot_bgcolor="white", paper_bgcolor="white",
                margin=dict(l=10, r=10, t=30, b=40),
                font=dict(size=12),
                xaxis=dict(tickformat="d"),
            )
            st.plotly_chart(fig, use_container_width=True)

        # ── SECTION 4: Final result ──────────────────────────
        st.markdown("---")
        st.markdown('<div class="section-title">Final IAS 19 Liability Calculation</div>',
                    unsafe_allow_html=True)
        st.markdown(render_result_banner(gross, emp["existing_fund"], net),
                    unsafe_allow_html=True)
        st.markdown(f"""
        <div style="background:white;border-radius:10px;padding:1rem 1.5rem;
                    box-shadow:0 1px 6px rgba(0,0,0,.07);font-size:.85rem">
          <table style="width:100%;border-collapse:collapse">
            <tr style="border-bottom:1px solid #f1f5f9">
              <td style="padding:.4rem .5rem;color:#64748b;font-weight:600">Step 1 — Gross DBO</td>
              <td style="padding:.4rem .5rem;color:#1e293b">
                Sum of all discounted expected severance payments across all projection years
              </td>
              <td style="padding:.4rem .5rem;font-weight:700;color:#7c3aed;text-align:right">
                ₪ {gross:,.0f}
              </td>
            </tr>
            <tr style="border-bottom:1px solid #f1f5f9">
              <td style="padding:.4rem .5rem;color:#64748b;font-weight:600">Step 2 — Plan Assets</td>
              <td style="padding:.4rem .5rem;color:#1e293b">
                Existing severance fund balance (deducted as plan assets under IAS 19)
              </td>
              <td style="padding:.4rem .5rem;font-weight:700;color:#0891b2;text-align:right">
                − ₪ {emp['existing_fund']:,.0f}
              </td>
            </tr>
            <tr>
              <td style="padding:.4rem .5rem;color:#1e3a5f;font-weight:700;font-size:.92rem">
                Net Liability (IAS 19)
              </td>
              <td style="padding:.4rem .5rem;color:#64748b;font-size:.82rem">
                Recognised on the balance sheet as a defined benefit liability
              </td>
              <td style="padding:.4rem .5rem;font-weight:800;color:#dc2626;
                          font-size:1.05rem;text-align:right">
                ₪ {net:,.0f}
              </td>
            </tr>
          </table>
        </div>
        """, unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════
    # TAB 3 – Charts
    # ══════════════════════════════════════════════════════════
    with tab_charts:
        active_df = df[df["active"] & ~df["left_2024"]].copy()
        active_df["full_name"] = active_df["first_name"] + " " + active_df["last_name"]

        r1, r2 = st.columns(2, gap="large")

        with r1:
            st.markdown('<div class="section-title">Top 15 Employees by Net Liability</div>',
                        unsafe_allow_html=True)
            top15 = (active_df.nlargest(15, "net_liability")
                              [["full_name", "net_liability"]])
            fig2 = px.bar(top15, x="net_liability", y="full_name",
                          orientation="h",
                          text=top15["net_liability"].map("₪{:,.0f}".format),
                          color_discrete_sequence=["#1d4ed8"])
            fig2.update_traces(textposition="outside", cliponaxis=False)
            fig2.update_layout(
                height=420, showlegend=False,
                xaxis_title="Net Liability (₪)", yaxis_title="",
                plot_bgcolor="white", paper_bgcolor="white",
                margin=dict(l=10, r=60, t=10, b=40),
                font=dict(size=11),
            )
            st.plotly_chart(fig2, use_container_width=True)

        with r2:
            st.markdown('<div class="section-title">Liability Distribution by Section 14 Coverage</div>',
                        unsafe_allow_html=True)
            def sec14_cat(p):
                if p == 1.0:   return "100% – Full coverage"
                if p >= 0.70:  return "72% – Partial"
                if p > 0:      return "Other partial"
                return "No Section 14"
            active_df["sec14_cat"] = active_df["sec14_pct"].apply(sec14_cat)
            pie_data = (active_df.groupby("sec14_cat")["net_liability"]
                                 .sum().reset_index())
            fig3 = px.pie(pie_data, values="net_liability", names="sec14_cat",
                          color_discrete_sequence=["#1d4ed8", "#0891b2",
                                                   "#6d28d9", "#94a3b8"],
                          hole=0.45)
            fig3.update_traces(
                textinfo="percent+label",
                hovertemplate="%{label}<br>₪%{value:,.0f}<extra></extra>")
            fig3.update_layout(
                height=380, showlegend=False,
                margin=dict(l=20, r=20, t=20, b=20))
            st.plotly_chart(fig3, use_container_width=True)

        st.markdown('<div class="section-title">Net Liability by Age Group</div>',
                    unsafe_allow_html=True)
        def age_bucket(a):
            if a < 30:   return "18–29"
            elif a < 40: return "30–39"
            elif a < 50: return "40–49"
            elif a < 60: return "50–59"
            else:        return "60+"
        active_df["age_group"] = active_df["age"].apply(age_bucket)
        age_grp = (active_df.groupby("age_group")
                             .agg(Count=("full_name","count"),
                                  Net_Liability=("net_liability","sum"))
                             .reset_index())
        fig4 = px.bar(
            age_grp, x="age_group", y="Net_Liability",
            text=age_grp["Net_Liability"].map("₪{:,.0f}".format),
            color_discrete_sequence=["#1d4ed8"],
            labels={"age_group": "Age Group", "Net_Liability": "Net Liability (₪)"},
        )
        fig4.update_traces(textposition="outside")
        fig4.update_layout(
            height=320,
            plot_bgcolor="white", paper_bgcolor="white",
            margin=dict(l=10, r=10, t=20, b=40),
            font=dict(size=12),
        )
        st.plotly_chart(fig4, use_container_width=True)


if __name__ == "__main__":
    main()
